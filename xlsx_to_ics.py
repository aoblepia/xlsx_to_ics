import sys
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QDialog, QLineEdit, QPushButton, QFileDialog, QVBoxLayout, QHBoxLayout, QComboBox, QMessageBox
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from dateutil import parser
from icalendar import Calendar, Event
from itertools import zip_longest
import pytz

class ExcelToICalConverter(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Excel to iCal Converter')
        self.setGeometry(100, 100, 500, 250)

        self.input_label = QLabel('Select Excel Spreadsheet:')
        self.input_line_edit = QLineEdit(self)
        self.browse_input_button = QPushButton('Browse', self)
        self.browse_input_button.clicked.connect(self.browse_input)

        self.output_label = QLabel('Select iCalendar File:')
        self.output_line_edit = QLineEdit(self)
        self.browse_output_button = QPushButton('Browse', self)
        self.browse_output_button.clicked.connect(self.browse_output)

        self.configure_button = QPushButton('Configure', self)
        self.configure_button.clicked.connect(self.match_columns)

        self.time_zone_label = QLabel('Select Default Time Zone:')
        self.time_zone_combobox = QComboBox(self)
        self.populate_time_zones()

        self.convert_button = QPushButton('Convert', self)
        self.convert_button.clicked.connect(self.convert)

        self.close_button = QPushButton('Close', self)
        self.close_button.clicked.connect(self.close_app)

        self.column_mapping = self.match_columns()

        input_layout = QHBoxLayout()
        input_layout.addWidget(self.input_line_edit)
        input_layout.addWidget(self.browse_input_button)

        output_layout = QHBoxLayout()
        output_layout.addWidget(self.output_line_edit)
        output_layout.addWidget(self.browse_output_button)

        time_zone_layout = QHBoxLayout()
        time_zone_layout.addWidget(self.time_zone_combobox)

        button_layout = QHBoxLayout()
        #close on left, convert on right
        button_layout.addWidget(self.close_button)
        button_layout.addWidget(self.convert_button)
        

        vbox = QVBoxLayout()
        vbox.addWidget(self.input_label)
        vbox.addLayout(input_layout)

        vbox.addWidget(self.output_label)
        vbox.addLayout(output_layout)

        vbox.addWidget(self.configure_button)

        vbox.addWidget(self.time_zone_label)
        vbox.addLayout(time_zone_layout)

        vbox.addLayout(button_layout)

        self.setLayout(vbox)

    def browse_input(self):
        file_dialog = QFileDialog()
        input_file, _ = file_dialog.getOpenFileName(self, 'Select Excel Spreadsheet')
        if input_file:
            self.input_line_edit.setText(input_file)
            

    def browse_output(self):
        file_dialog = QFileDialog()
        output_location, _ = file_dialog.getSaveFileName(self, 'Save iCalendar File', filter='*.ics')
        if output_location:
            self.output_line_edit.setText(output_location)

    def populate_time_zones(self):
        time_zones = pytz.all_timezones
        self.time_zone_combobox.addItems(time_zones)

        # Set default time zone to "US/Eastern"
        default_time_zone = "US/Eastern"
        default_index = self.time_zone_combobox.findText(default_time_zone)
        if default_index != -1:
            self.time_zone_combobox.setCurrentIndex(default_index)

    def match_columns(self):
        self.criteria = ["event_summary", "start_date", "start_time", "end_date", "end_time", "Ignore"]  # Add your criteria here
        

        input_file = self.input_line_edit.text()  
        if input_file:
            workbook = load_workbook(input_file)
            sheet = workbook.active

            # Assuming the first row contains the column headers
            headers = [cell.column for cell in sheet[1]]

            # Create a mapping between columns and criteria
            column_mapping = {}
            #criteria_dialog = CriteriaDialog(headers, self.criteria, self.column_mapping)
            criteria_dialog = CriteriaDialog(headers, self.criteria)
            result = criteria_dialog.exec_()

            if result == QDialog.Accepted:
                column_mapping = criteria_dialog.get_column_mapping()

                #print("Column Mapping:")
                #for column, criteria in column_mapping.items():
                #    print(f"{column}: {criteria}")
                return column_mapping
            
            workbook.close()
        else:
            #print("Please select an Excel file first.")
            print("Default config loaded.")
            column_mapping = {
                1 : "event_summary",
                2 : "start_date",
                3 : "start_time",
                4 : "end_date",
                5 : "end_time"
            }
            return column_mapping

    def convert(self):
        input_file = self.input_line_edit.text()
        output_location = self.output_line_edit.text()
        selected_time_zone = self.time_zone_combobox.currentText()

        if not input_file or not output_location:
            QMessageBox.warning(self, 'Warning', 'Please select both input and output files.')
            return

        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active

        cal = Calendar()

        selected_tz = pytz.timezone(selected_time_zone)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # ensure that the sheet has the minimum number of rows, even if data is excluded.
            while len(row) < 5:
                row += ("",)
            # for each element in dictionary, match key (based on value) to criteria
            for key, value in self.column_mapping.items():
                match value:
                    case "event_summary":
                        summary = row[key-1]
                    case "start_date":
                        start_date = row[key-1]
                    case "start_time":
                        start_time = row[key-1]
                    case "end_date":
                        end_date = row[key-1]
                    case "end_time":
                        end_time = row[key-1]
                    case _:
                        break

            event = Event()
            event.add('summary', summary)

            # Combine date and time strings, then parse with dateutil.parser
            start_datetime = parser.parse(f"{start_date} {start_time}")
            end_datetime = parser.parse(f"{end_date} {end_time}")

            # Set time zone
            start_datetime = selected_tz.localize(start_datetime)
            end_datetime = selected_tz.localize(end_datetime)

            event.add('dtstart', start_datetime)
            event.add('dtend', end_datetime)

            cal.add_component(event)

        if output_location:
            with open(output_location, 'wb') as f:
                f.write(cal.to_ical())
            QMessageBox.information(self, 'Conversion Completed', f'Conversion completed. Output file: {output_location}')
        else:
            QMessageBox.warning(self, 'Warning', 'No output location selected.')

    def close_app(self):
        self.close()

class CriteriaDialog(QDialog):
    #def __init__(self, column_names, criteria_list, column_mapping):
    def __init__(self, column_names, criteria_list):
        super().__init__()

        self.column_names = column_names
        self.criteria_list = criteria_list
        #self.column_mapping = column_mapping
        self.column_mapping = {}

        self.init_ui()

    def init_ui(self):
        self.layout = QVBoxLayout()

        self.criteria_widgets = []

        #for k1, k2 in zip_longest(self.column_names, self.column_mapping.values(), fillvalue='Ignore'):
        #    criteria_widget = ColumnCriteriaWidget(k1, self.column_mapping)
        #    self.layout.addWidget(criteria_widget)
        #    self.criteria_widgets.append(criteria_widget)

        for column_name in self.column_names:
            criteria_widget = ColumnCriteriaWidget(column_name, self.criteria_list)
            self.layout.addWidget(criteria_widget)
            self.criteria_widgets.append(criteria_widget)

        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

        self.setLayout(self.layout)

    def get_column_mapping(self):
        for widget in self.criteria_widgets:
            self.column_mapping[widget.column_number] = widget.get_selected_criteria()
        return self.column_mapping

class ColumnCriteriaWidget(QWidget):
    def __init__(self, column_number, criteria_list):
        super().__init__()

        self.column_number = column_number
        self.criteria_list = criteria_list
        self.selected_criteria = None

        self.init_ui()

    def init_ui(self):
        self.layout = QHBoxLayout()

        self.label = QLabel(f"Select criteria for column {self.column_number}:")
        self.layout.addWidget(self.label)

        self.criteria_combo = QComboBox()
        self.criteria_combo.addItems(self.criteria_list)
        self.layout.addWidget(self.criteria_combo)

        self.setLayout(self.layout)

    def get_selected_criteria(self):
        return self.criteria_combo.currentText()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    converter = ExcelToICalConverter()
    converter.show()
    sys.exit(app.exec_())
